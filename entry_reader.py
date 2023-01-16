import pandas as pd
import datetime
import sys
from tempfile import TemporaryFile
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side

st_dt = sys.argv[1]
ed_dt = sys.argv[2]
date_list = pd.date_range(start=st_dt, end=ed_dt, freq='D')
fp = TemporaryFile('w+t')
# txt 파일생성
# fp = open('test.txt', 'w')

for x in date_list:
    fileName = f'entry.{str(x)[:10]}.log'
    try:
        with open(f'./entry/{fileName}', 'r', encoding='UTF-8') as f:
            fileContent = f.readlines()

        fileContent = [x.strip() for x in fileContent]
        for line in fileContent:
            splitLine = line.split()
            if 'hdr' in splitLine:
                fp.write(' '.join(splitLine))
            elif 'bdy' in splitLine:
                fp.write(' '.join(splitLine))
            elif 'jsonResult' in splitLine:
                fp.write(' '.join(splitLine))
                fp.write('\n')
    except:
        print(f'{str(x)[:10]} 파일없음')

# txt 파일생성
# fp.close

# 엑셀
write_wb = Workbook()
# Sheet1에다 입력
write_ws = write_wb.active
write_ws.column_dimensions['B'].width = 14.5
write_ws.column_dimensions['C'].width = 14.5
write_ws.column_dimensions['D'].width = 14.5
write_ws['B2'] = '날짜'
write_ws['C2'] = '시분초'
write_ws['D2'] = '고객명'
write_ws['B2'].fill = PatternFill(
    start_color='D9E1F2', fill_type='solid')  # 배경색 추가
write_ws['C2'].fill = PatternFill(
    start_color='D9E1F2', fill_type='solid')  # 배경색 추가
write_ws['D2'].fill = PatternFill(
    start_color='D9E1F2', fill_type='solid')  # 배경색 추가

fp.seek(0)
fpContent = fp.readlines()
for x in fpContent:
    if 'STORE_CD=30507' in x:
        if '이미 사용된 코드 입니다' not in x:
            if 'QR 코드가 정확하지 않습니다' not in x:
                if '정상적으로 생성한 코드가 아닙니다' not in x:
                    req_dt_index = x.find('REQ_DT=')
                    # 날짜
                    tran_ymd_st_index = req_dt_index + 7
                    tran_ymd_ed_index = tran_ymd_st_index + 8
                    tran_ymd = x[tran_ymd_st_index:tran_ymd_ed_index]
                    # 시분초
                    req_dt_st_index = req_dt_index + 15
                    req_dt_ed_index = req_dt_st_index + 6
                    req_dt = x[req_dt_st_index:req_dt_ed_index]
                    # 고객명
                    cust_nm_index = x.find('CUST_NM') + 12
                    cust_nm = x[cust_nm_index:]
                    cust_nm = cust_nm.strip().replace('"', '')
                    cust_nm = cust_nm.replace('}', '')
                    # 고객ID
                    # cust_id_index = x.find('CUST_ID') + 12
                    # cust_id_ed_index = cust_id_index + 9
                    # cust_id = x[cust_id_index:cust_id_ed_index]
                    # cust_id = cust_id.strip().replace('"', '')
                    # write_ws.append([cust_id, tran_ymd, req_dt, cust_nm, x])
                    write_ws.append(['', tran_ymd, req_dt, cust_nm])
                    # print(tran_ymd, ', ', req_dt, ', ', cust_nm)

columns = ['B', 'C', 'D']

for x in range(2, write_ws.max_row+1):
    for y in columns:
        write_ws[f'{y}{x}'].border = Border(left=Side(style="thin", color="000000"),
                                            right=Side(
            style='thin', color="000000"),
            top=Side(
            style="thin", color="000000"),
            bottom=Side(
            style="thin", color="000000"),
            diagonal=Side(
            style="thin", color="000000"),
            diagonal_direction=0,
            outline=Side(
            style="thin", color="000000"),
            vertical=Side(
            style="thin", color="000000"),
            horizontal=Side(style="thin", color="000000"))

fp.close
write_wb.save(
    f'./result_entry/무인점포_입장이력_삼성병원2호점_{(datetime.date.today().strftime("%Y%m%d"))}.xlsx')
print((f'무인점포_입장이력_삼성병원2호점_{(datetime.date.today().strftime("%Y%m%d"))}.xlsx'))
