import pandas as pd
import datetime
import sys
from tempfile import TemporaryFile
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

st_dt = sys.argv[1]
ed_dt = sys.argv[2]
date_list = pd.date_range(start=st_dt, end=ed_dt, freq='D')
fp = TemporaryFile('w+t')
# txt 파일생성
# fp = open('test.txt', 'w')

for x in date_list:
    fileName = f'tran_rcv.{str(x)[:10]}.log'
    try:
        with open(f'./tran_rcv/{fileName}', 'r', encoding='UTF-8') as f:
            fileContent = f.readlines()

        fileContent = [x.strip() for x in fileContent]
        for line in fileContent:
            splitLine = line.split()
            if '[RCV]' in splitLine:
                fp.write(' '.join(splitLine))
            elif 'hdr' in splitLine:
                fp.write(' '.join(splitLine))
                fp.write('\n')
            elif 'data.' in splitLine:
                fp.write('\n')
    except:
        print(f'{str(x)[:10]} 파일없음')

# txt 파일생성
# fp.close


# 엑셀
write_wb = Workbook()
# Sheet1에다 입력
write_ws = write_wb.active
write_ws.column_dimensions['B'].width = 14.5
write_ws.column_dimensions['C'].width = 14.5
write_ws.column_dimensions['D'].width = 14.5
write_ws['B2'] = '날짜'
write_ws['C2'] = '시분초'
write_ws['D2'] = '고객명'
write_ws['B2'].fill = PatternFill(
    start_color='D9E1F2', fill_type='solid')  # 배경색 추가
write_ws['C2'].fill = PatternFill(
    start_color='D9E1F2', fill_type='solid')  # 배경색 추가
write_ws['D2'].fill = PatternFill(
    start_color='D9E1F2', fill_type='solid')  # 배경색 추가

fp.seek(0)
fpContent = fp.readlines()
row_cnt = 2

for x in fpContent:
    if 'STORE_CD=30507' in x:
        row_cnt = row_cnt + 1
        req_dt_index = x.find('SYS_YMDHMS')
        # 날짜
        tran_ymd_st_index = req_dt_index + 13
        tran_ymd_ed_index = tran_ymd_st_index + 8
        tran_ymd = x[tran_ymd_st_index:tran_ymd_ed_index]
        # 시분초
        req_dt_st_index = req_dt_index + 21
        req_dt_ed_index = req_dt_st_index + 6
        req_dt = x[req_dt_st_index:req_dt_ed_index]
        # 고객명
        cust_no_index = x.find('MOBILE_ID=') + 10
        cust_no = x[cust_no_index:]
        cust_no = cust_no.strip().replace('"', '')
        cust_no = cust_no.replace('}', '')
        write_ws.append(
            ['', tran_ymd, req_dt, f'=VLOOKUP(E{row_cnt}, Sheet2!A2:B41781, 2, FALSE)', cust_no])

write_ws.column_dimensions['E'].hidden = True

columns = ['B', 'C', 'D']

for x in range(2, write_ws.max_row+1):
    for y in columns:
        write_ws[f'{y}{x}'].border = Border(left=Side(style="thin", color="000000"),
                                            right=Side(
            style='thin', color="000000"),
            top=Side(
            style="thin", color="000000"),
            bottom=Side(
            style="thin", color="000000"),
            diagonal=Side(
            style="thin", color="000000"),
            diagonal_direction=0,
            outline=Side(
            style="thin", color="000000"),
            vertical=Side(
            style="thin", color="000000"),
            horizontal=Side(style="thin", color="000000"))

fp.close

write_ws2 = write_wb.create_sheet('Sheet2', 1)

load_tmp_xlsx = load_workbook('./tmp.xlsx')
tmp_sheet = load_tmp_xlsx['Sheet2']
for row in tmp_sheet:
    write_ws2.append([row[0].value, row[1].value, row[2].value, row[3].value])

write_wb.save(
    f'./result_tran/무인점포_구매이력_삼성병원2호점_{(datetime.date.today().strftime("%Y%m%d"))}.xlsx')
print((f'무인점포_구매이력_삼성병원2호점_{(datetime.date.today().strftime("%Y%m%d"))}.xlsx'))
